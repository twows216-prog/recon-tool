"""
Card Reconciliation Tool - Web Version
Host on Streamlit Cloud for free: https://streamlit.io/cloud
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


# ============== PAGE CONFIG ==============
st.set_page_config(
    page_title="Card Reconciliation Tool",
    page_icon="💳",
    layout="wide"
)

# ============== STYLES ==============
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1E88E5;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .card-box {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .success-box {
        background-color: #d4edda;
        border-radius: 10px;
        padding: 1rem;
        border-left: 5px solid #28a745;
    }
    .error-box {
        background-color: #f8d7da;
        border-radius: 10px;
        padding: 1rem;
        border-left: 5px solid #dc3545;
    }
</style>
""", unsafe_allow_html=True)


# ============== FUNCTIONS ==============
def extract_card_number(filename):
    """Extract card number from filename - very flexible matching"""
    filename_lower = filename.lower()
    
    # P2P Statement pattern: contains "statement" and "d17-XXX-"
    if 'statement' in filename_lower:
        match = re.search(r'd17-(\d+)-', filename_lower)
        if match:
            return str(int(match.group(1))), 'p2p_statement'
    
    # Card Journal patterns - anything that starts with a number
    # Examples: "342-journal-31-12-2025.csv", "308-d17-journal.csv", "075-D17-31-01-2026.csv"
    match = re.match(r'^(\d+)', filename_lower)
    if match:
        return str(int(match.group(1))), 'card_journal'
    
    return None, None


def load_card_journal(file):
    """Load and parse the Card Journal CSV"""
    try:
        df = pd.read_csv(file, sep=';')
        if len(df.columns) == 1:
            file.seek(0)
            df = pd.read_csv(file, sep=',')
    except:
        file.seek(0)
        df = pd.read_csv(file, sep=',')
    
    df.columns = df.columns.str.strip()
    
    if 'DATE' in df.columns:
        df['DATE'] = df['DATE'].astype(str).str.strip()
        try:
            df['DATE'] = pd.to_datetime(df['DATE'], format='%d/%m/%Y')
        except:
            df['DATE'] = pd.to_datetime(df['DATE'], dayfirst=True)
    
    if 'MONTANT' in df.columns:
        df['MONTANT'] = df['MONTANT'].astype(str).str.replace(',', '.').astype(float)
    
    if 'LIBELLE' in df.columns:
        df['Auth'] = df['LIBELLE'].str.extract(r'(\d{6})$')
    
    return df


def load_p2p_statement(file):
    """Load and parse the P2P Statement CSV"""
    df = pd.read_csv(file)
    df.columns = df.columns.str.strip()
    
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'])
    
    # Convert Auth to string (handle NaN and float)
    if 'Auth' in df.columns:
        df['Auth'] = df['Auth'].fillna('').astype(str)
        df['Auth'] = df['Auth'].str.replace(r'\.0$', '', regex=True)
    
    return df


def get_in_transactions_card(df, from_date):
    """Get incoming transactions from Card Journal (Transfert du)"""
    filtered = df[df['DATE'] >= from_date].copy()
    if 'LIBELLE' not in filtered.columns or filtered.empty:
        return pd.DataFrame()
    
    incoming = filtered[filtered['LIBELLE'].str.contains('Transfert du', case=False, na=False)].copy()
    
    if incoming.empty:
        return pd.DataFrame()
    
    incoming = incoming[incoming['MONTANT'] > 0].copy()
    
    if incoming.empty:
        return pd.DataFrame()
    
    incoming['Auth'] = incoming['Auth'].astype(str)
    return incoming.reset_index(drop=True)


def get_out_transactions_card(df, from_date, include_withdrawal=True, include_mandat=True):
    """Get outgoing transactions from Card Journal (Transfert vers + Mandat)"""
    filtered = df[df['DATE'] >= from_date].copy()
    if 'LIBELLE' not in filtered.columns or filtered.empty:
        return pd.DataFrame()
    
    # Build filter conditions based on options
    conditions = pd.Series([False] * len(filtered), index=filtered.index)
    
    if include_withdrawal:
        conditions = conditions | filtered['LIBELLE'].str.contains('Transfert vers', case=False, na=False)
    
    if include_mandat:
        conditions = conditions | filtered['LIBELLE'].str.contains('Mandat', case=False, na=False)
    
    outgoing = filtered[conditions].copy()
    
    if outgoing.empty:
        return pd.DataFrame()
    
    outgoing['MONTANT'] = outgoing['MONTANT'].abs()  # Make positive for comparison
    
    # Extract account number for "Transfert vers" (e.g., "Transfert vers 29526566 230314" -> "29526566")
    outgoing['ToAccount'] = outgoing['LIBELLE'].apply(
        lambda x: re.search(r'Transfert vers (\d+)', str(x)).group(1) if re.search(r'Transfert vers (\d+)', str(x)) else None
    )
    
    # For Mandat, we'll match by amount and date only
    outgoing['IsMandat'] = outgoing['LIBELLE'].str.contains('Mandat', case=False, na=False)
    
    # Create a match key: Date + Amount + ToAccount (for transfers) or Date + Amount + MANDAT (for mandats)
    def create_match_key(row):
        date_str = row['DATE'].strftime('%Y-%m-%d')
        amount_str = f"{row['MONTANT']:.2f}"
        if pd.notna(row['ToAccount']) and row['ToAccount']:
            return f"{date_str}_{amount_str}_{row['ToAccount']}"
        else:
            return f"{date_str}_{amount_str}_MANDAT"
    
    outgoing['MatchKey'] = outgoing.apply(create_match_key, axis=1)
    
    return outgoing.reset_index(drop=True)


def get_in_transactions_p2p(df, from_date):
    """Get incoming transactions from P2P Statement (DEPOSIT)"""
    if 'Type' in df.columns:
        deposits = df[df['Type'] == 'DEPOSIT'].copy()
    else:
        deposits = df.copy()
    
    if deposits.empty:
        return pd.DataFrame()
    
    # Apply date filter
    if 'Date' in deposits.columns:
        deposits = deposits[deposits['Date'] >= from_date].copy()
    
    if deposits.empty:
        return pd.DataFrame()
    
    deposits = deposits[deposits['Auth'] != ''].copy()
    
    if deposits.empty:
        return pd.DataFrame()
    
    if 'Amount' in deposits.columns:
        deposits['Adjusted_Amount'] = deposits['Amount'].apply(
            lambda x: x * 0.99 if x > 40 else x
        )
    
    deposits['Auth'] = deposits['Auth'].astype(str)
    return deposits.reset_index(drop=True)


def get_out_transactions_p2p(df, from_date, include_withdrawal=True, include_cashout=True):
    """Get outgoing transactions from P2P Statement (WITHDRAWAL + CASHOUT)"""
    if 'Type' not in df.columns:
        return pd.DataFrame()
    
    # Build filter conditions based on options
    conditions = pd.Series([False] * len(df), index=df.index)
    
    if include_withdrawal:
        conditions = conditions | (df['Type'] == 'WITHDRAWAL')
    
    if include_cashout:
        conditions = conditions | (df['Type'] == 'CASHOUT')
    
    withdrawals = df[conditions].copy()
    
    if withdrawals.empty:
        return pd.DataFrame()
    
    # Apply date filter
    if 'Date' in withdrawals.columns:
        withdrawals = withdrawals[withdrawals['Date'] >= from_date].copy()
    
    if withdrawals.empty:
        return pd.DataFrame()
    
    if 'Amount' in withdrawals.columns:
        withdrawals['Amount'] = withdrawals['Amount'].abs()  # Make positive for comparison
        withdrawals['Adjusted_Amount'] = withdrawals['Amount']  # No fee adjustment for withdrawals
    
    # Get To Account for WITHDRAWAL, empty for CASHOUT
    if 'To Account' in withdrawals.columns:
        withdrawals['ToAccount'] = withdrawals['To Account'].apply(
            lambda x: str(int(x)) if pd.notna(x) and x != '' else ''
        )
    else:
        withdrawals['ToAccount'] = ''
    
    # Create match key: Date + Amount + ToAccount (for WITHDRAWAL) or Date + Amount + MANDAT (for CASHOUT)
    def create_match_key(row):
        date_str = row['Date'].strftime('%Y-%m-%d')
        amount_str = f"{row['Amount']:.2f}"
        if row['ToAccount'] and row['ToAccount'] != '':
            return f"{date_str}_{amount_str}_{row['ToAccount']}"
        else:
            return f"{date_str}_{amount_str}_MANDAT"
    
    withdrawals['MatchKey'] = withdrawals.apply(create_match_key, axis=1)
    
    return withdrawals.reset_index(drop=True)


def reconcile_transactions(card_df, p2p_df, direction='IN'):
    """Reconcile transactions between Card Journal and P2P Statement"""
    
    if card_df.empty and p2p_df.empty:
        return {
            'summary': {'card_count': 0, 'p2p_count': 0, 'matched': 0, 
                       'missing_in_p2p': 0, 'missing_in_card': 0},
            'missing_in_p2p': [], 'missing_in_card': [], 
            'matched_details': [], 'discrepancies': [],
            'totals': {'card': 0, 'p2p': 0, 'adjusted': 0, 'difference': 0},
            'missing_amount': 0
        }
    
    # For IN transactions, use Auth code; for OUT, use MatchKey
    if direction == 'IN':
        card_keys = set(card_df['Auth'].dropna()) if not card_df.empty else set()
        p2p_keys = set(p2p_df['Auth'].dropna()) if not p2p_df.empty else set()
        key_field_card = 'Auth'
        key_field_p2p = 'Auth'
    else:  # OUT
        card_keys = set(card_df['MatchKey'].dropna()) if not card_df.empty and 'MatchKey' in card_df.columns else set()
        p2p_keys = set(p2p_df['MatchKey'].dropna()) if not p2p_df.empty and 'MatchKey' in p2p_df.columns else set()
        key_field_card = 'MatchKey'
        key_field_p2p = 'MatchKey'
    
    missing_in_p2p = card_keys - p2p_keys
    missing_in_card = p2p_keys - card_keys
    matched = card_keys & p2p_keys
    
    results = {
        'summary': {
            'card_count': len(card_df),
            'p2p_count': len(p2p_df),
            'matched': len(matched),
            'missing_in_p2p': len(missing_in_p2p),
            'missing_in_card': len(missing_in_card)
        },
        'missing_in_p2p': [],
        'missing_in_card': [],
        'matched_details': [],
        'discrepancies': []
    }
    
    # Missing in P2P
    for key in missing_in_p2p:
        if not card_df.empty:
            rows = card_df[card_df[key_field_card] == key]
            if not rows.empty:
                row = rows.iloc[0]
                results['missing_in_p2p'].append({
                    'Auth': row.get('Auth', key) if direction == 'IN' else row.get('ToAccount', '-'),
                    'Date': row['DATE'].strftime('%d/%m/%Y'),
                    'Time': str(row.get('HEURE', '')).strip(),
                    'Amount': abs(row['MONTANT']),
                    'Description': row.get('LIBELLE', '')
                })
    
    # Missing in Card
    for key in missing_in_card:
        if not p2p_df.empty:
            rows = p2p_df[p2p_df[key_field_p2p] == key]
            if not rows.empty:
                row = rows.iloc[0]
                results['missing_in_card'].append({
                    'Auth': row.get('Auth', '-') if direction == 'IN' else row.get('ToAccount', '-'),
                    'Date': row['Date'].strftime('%d/%m/%Y'),
                    'Time': str(row.get('Time', '')),
                    'Amount': row['Amount'],
                    'Type': row.get('Type', '-')
                })
    
    # Matched details
    total_card = 0
    total_p2p = 0
    total_adjusted = 0
    
    for key in sorted(matched):
        card_rows = card_df[card_df[key_field_card] == key]
        p2p_rows = p2p_df[p2p_df[key_field_p2p] == key]
        
        if not card_rows.empty and not p2p_rows.empty:
            card_row = card_rows.iloc[0]
            p2p_row = p2p_rows.iloc[0]
            
            card_amt = abs(card_row['MONTANT'])
            p2p_amt = p2p_row['Amount']
            adj_amt = p2p_row['Adjusted_Amount']
            diff = round(card_amt - adj_amt, 2)
            
            total_card += card_amt
            total_p2p += p2p_amt
            total_adjusted += adj_amt
            
            status = "OK" if abs(diff) < 0.01 else "DIFF"
            fee_applied = "Yes" if p2p_amt > 40 and direction == 'IN' else "No"
            
            detail = {
                'Auth': card_row.get('Auth', '-') if direction == 'IN' else card_row.get('ToAccount', '-'),
                'Date': card_row['DATE'].strftime('%d/%m/%Y'),
                'P2P_Amount': p2p_amt,
                'Adjusted_Amount': round(adj_amt, 2),
                'Card_Amount': card_amt,
                'Difference': diff,
                'Status': status,
                'Fee_Applied': fee_applied
            }
            results['matched_details'].append(detail)
            
            if abs(diff) >= 0.01:
                results['discrepancies'].append(detail)
    
    results['totals'] = {
        'card': round(total_card, 2),
        'p2p': round(total_p2p, 2),
        'adjusted': round(total_adjusted, 2),
        'difference': round(total_card - total_adjusted, 2)
    }
    
    if results['missing_in_p2p']:
        results['missing_amount'] = sum(item['Amount'] for item in results['missing_in_p2p'])
    else:
        results['missing_amount'] = 0
    
    return results


def create_excel_report(all_results, from_date, include_withdrawal, include_mandat, include_cashout):
    """Create professional Excel report with summary and details"""
    wb = Workbook()
    
    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill_blue = PatternFill("solid", fgColor="4472C4")
    header_fill_green = PatternFill("solid", fgColor="548235")
    header_fill_purple = PatternFill("solid", fgColor="7030A0")
    header_fill_red = PatternFill("solid", fgColor="C00000")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    ok_font = Font(color="006100")
    error_fill = PatternFill("solid", fgColor="FFC7CE")
    error_font = Font(color="9C0006")
    warning_fill = PatternFill("solid", fgColor="FFEB9C")
    gray_fill = PatternFill("solid", fgColor="F2F2F2")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ==================== SUMMARY SHEET ====================
    ws = wb.active
    ws.title = "RECONCILIATION REPORT"
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "CARD RECONCILIATION REPORT"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Report Info
    ws['A3'] = "Report Date:"
    ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws['A4'] = "Period From:"
    ws['B4'] = from_date.strftime('%d/%m/%Y')
    ws['A5'] = "Options:"
    options_text = []
    if include_withdrawal:
        options_text.append("Withdrawal included")
    else:
        options_text.append("Withdrawal excluded")
    if include_mandat:
        options_text.append("Mandat included")
    else:
        options_text.append("Mandat excluded")
    if include_cashout:
        options_text.append("Cashout included")
    else:
        options_text.append("Cashout excluded")
    ws['B5'] = " | ".join(options_text)
    
    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)
    ws['A5'].font = Font(bold=True)
    
    # Rules
    ws['A7'] = "Matching Rules:"
    ws['A7'].font = Font(bold=True, size=11)
    ws['A8'] = "• IN (Deposits): Match by Auth code | 1% fee removed from amounts > 40"
    ws['A9'] = "• OUT (Withdrawals): Match by Date + Amount + Account | No fee adjustment"
    
    # ==================== SUMMARY TABLE ====================
    row = 11
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "📊 SUMMARY BY CARD"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = gray_fill
    row += 2
    
    # IN Summary Header
    ws[f'A{row}'] = "📥 IN TRANSACTIONS (Deposits)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="4472C4")
    row += 1
    
    in_headers = ['Card', 'Card Journal', 'P2P Statement', 'Matched', 'Missing P2P', 'Missing Card', 'Missing Amount', 'Status']
    for col, header in enumerate(in_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    total_in_missing = 0
    for card_num, results in sorted(all_results.items()):
        r = results['in']
        summary = r['summary']
        
        ws.cell(row=row, column=1, value=card_num).border = border
        ws.cell(row=row, column=2, value=summary['card_count']).border = border
        ws.cell(row=row, column=3, value=summary['p2p_count']).border = border
        ws.cell(row=row, column=4, value=summary['matched']).border = border
        
        cell_mp = ws.cell(row=row, column=5, value=summary['missing_in_p2p'])
        cell_mp.border = border
        if summary['missing_in_p2p'] > 0:
            cell_mp.fill = error_fill
            cell_mp.font = error_font
        
        cell_mc = ws.cell(row=row, column=6, value=summary['missing_in_card'])
        cell_mc.border = border
        if summary['missing_in_card'] > 0:
            cell_mc.fill = error_fill
            cell_mc.font = error_font
        
        cell_ma = ws.cell(row=row, column=7, value=r['missing_amount'])
        cell_ma.border = border
        cell_ma.number_format = '#,##0.00'
        if r['missing_amount'] > 0:
            cell_ma.fill = error_fill
            cell_ma.font = error_font
        
        # Status
        if summary['missing_in_p2p'] == 0 and summary['missing_in_card'] == 0:
            status_cell = ws.cell(row=row, column=8, value="✓ OK")
            status_cell.fill = ok_fill
            status_cell.font = ok_font
        else:
            status_cell = ws.cell(row=row, column=8, value="✗ Issues")
            status_cell.fill = error_fill
            status_cell.font = error_font
        status_cell.border = border
        
        total_in_missing += r['missing_amount']
        row += 1
    
    # IN Total row
    ws.cell(row=row, column=1, value="TOTAL IN").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_in_missing).font = Font(bold=True)
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    if total_in_missing > 0:
        ws.cell(row=row, column=7).fill = error_fill
    row += 2
    
    # OUT Summary Header
    ws[f'A{row}'] = "📤 OUT TRANSACTIONS (Withdrawals/Cashouts)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="7030A0")
    row += 1
    
    out_headers = ['Card', 'Card Journal', 'P2P Statement', 'Matched', 'Missing P2P', 'Missing Card', 'Missing Amount', 'Status']
    for col, header in enumerate(out_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_purple
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    total_out_missing = 0
    for card_num, results in sorted(all_results.items()):
        r = results['out']
        summary = r['summary']
        
        ws.cell(row=row, column=1, value=card_num).border = border
        ws.cell(row=row, column=2, value=summary['card_count']).border = border
        ws.cell(row=row, column=3, value=summary['p2p_count']).border = border
        ws.cell(row=row, column=4, value=summary['matched']).border = border
        
        cell_mp = ws.cell(row=row, column=5, value=summary['missing_in_p2p'])
        cell_mp.border = border
        if summary['missing_in_p2p'] > 0:
            cell_mp.fill = error_fill
            cell_mp.font = error_font
        
        cell_mc = ws.cell(row=row, column=6, value=summary['missing_in_card'])
        cell_mc.border = border
        if summary['missing_in_card'] > 0:
            cell_mc.fill = error_fill
            cell_mc.font = error_font
        
        cell_ma = ws.cell(row=row, column=7, value=r['missing_amount'])
        cell_ma.border = border
        cell_ma.number_format = '#,##0.00'
        if r['missing_amount'] > 0:
            cell_ma.fill = error_fill
            cell_ma.font = error_font
        
        # Status
        if summary['missing_in_p2p'] == 0 and summary['missing_in_card'] == 0:
            status_cell = ws.cell(row=row, column=8, value="✓ OK")
            status_cell.fill = ok_fill
            status_cell.font = ok_font
        else:
            status_cell = ws.cell(row=row, column=8, value="✗ Issues")
            status_cell.fill = error_fill
            status_cell.font = error_font
        status_cell.border = border
        
        total_out_missing += r['missing_amount']
        row += 1
    
    # OUT Total row
    ws.cell(row=row, column=1, value="TOTAL OUT").font = Font(bold=True)
    ws.cell(row=row, column=7, value=total_out_missing).font = Font(bold=True)
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    if total_out_missing > 0:
        ws.cell(row=row, column=7).fill = error_fill
    row += 2
    
    # Grand Total
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "⚠️ TOTAL MISSING AMOUNT (IN + OUT):"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.cell(row=row, column=7, value=total_in_missing + total_out_missing)
    ws.cell(row=row, column=7).font = Font(bold=True, size=12)
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    if (total_in_missing + total_out_missing) > 0:
        ws.cell(row=row, column=7).fill = error_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 12
    
    # ==================== DETAIL SHEETS PER CARD ====================
    for card_num, results in sorted(all_results.items()):
        ws_card = wb.create_sheet(f"Card {card_num}")
        
        # Title
        ws_card.merge_cells('A1:H1')
        ws_card[f'A1'] = f"CARD {card_num} - RECONCILIATION DETAILS"
        ws_card['A1'].font = title_font
        ws_card['A1'].fill = title_fill
        ws_card['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_card.row_dimensions[1].height = 30
        
        row = 3
        
        # ===== IN SECTION =====
        ws_card.merge_cells(f'A{row}:H{row}')
        ws_card[f'A{row}'] = "📥 IN TRANSACTIONS (Deposits)"
        ws_card[f'A{row}'].font = Font(bold=True, size=14, color="4472C4")
        ws_card[f'A{row}'].fill = gray_fill
        row += 2
        
        r_in = results['in']
        
        # IN Summary
        ws_card[f'A{row}'] = "Card Journal:"
        ws_card[f'B{row}'] = r_in['summary']['card_count']
        ws_card[f'C{row}'] = "P2P Statement:"
        ws_card[f'D{row}'] = r_in['summary']['p2p_count']
        ws_card[f'E{row}'] = "Matched:"
        ws_card[f'F{row}'] = r_in['summary']['matched']
        ws_card[f'A{row}'].font = Font(bold=True)
        ws_card[f'C{row}'].font = Font(bold=True)
        ws_card[f'E{row}'].font = Font(bold=True)
        row += 2
        
        # Missing in P2P
        if r_in['missing_in_p2p']:
            ws_card[f'A{row}'] = "❌ MISSING IN P2P STATEMENT"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="C00000")
            row += 1
            
            headers = ['Auth', 'Date', 'Time', 'Amount', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_red
                cell.border = border
            row += 1
            
            for item in r_in['missing_in_p2p']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['Time']).border = border
                cell_amt = ws_card.cell(row=row, column=4, value=item['Amount'])
                cell_amt.border = border
                cell_amt.number_format = '#,##0.00'
                ws_card.cell(row=row, column=5, value=item['Description']).border = border
                for c in range(1, 6):
                    ws_card.cell(row=row, column=c).fill = error_fill
                row += 1
            
            ws_card[f'A{row}'] = "Total Missing:"
            ws_card[f'A{row}'].font = Font(bold=True)
            ws_card.cell(row=row, column=4, value=r_in['missing_amount'])
            ws_card.cell(row=row, column=4).font = Font(bold=True)
            ws_card.cell(row=row, column=4).number_format = '#,##0.00'
            row += 2
        
        # Missing in Card
        if r_in['missing_in_card']:
            ws_card[f'A{row}'] = "❌ MISSING IN CARD JOURNAL"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="C00000")
            row += 1
            
            headers = ['Auth', 'Date', 'Time', 'Amount', 'Type']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_red
                cell.border = border
            row += 1
            
            for item in r_in['missing_in_card']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['Time']).border = border
                cell_amt = ws_card.cell(row=row, column=4, value=item['Amount'])
                cell_amt.border = border
                cell_amt.number_format = '#,##0.00'
                ws_card.cell(row=row, column=5, value=item.get('Type', '-')).border = border
                for c in range(1, 6):
                    ws_card.cell(row=row, column=c).fill = error_fill
                row += 1
            row += 1
        
        # Matched IN
        if r_in['matched_details']:
            ws_card[f'A{row}'] = "✓ MATCHED TRANSACTIONS"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="006100")
            row += 1
            
            headers = ['Auth', 'Date', 'P2P Amount', 'Adjusted', 'Card Amount', 'Diff', 'Status', 'Fee']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_green
                cell.border = border
            row += 1
            
            for item in r_in['matched_details']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['P2P_Amount']).border = border
                ws_card.cell(row=row, column=4, value=item['Adjusted_Amount']).border = border
                ws_card.cell(row=row, column=5, value=item['Card_Amount']).border = border
                ws_card.cell(row=row, column=6, value=item['Difference']).border = border
                
                status_cell = ws_card.cell(row=row, column=7, value=item['Status'])
                status_cell.border = border
                if item['Status'] == 'OK':
                    status_cell.fill = ok_fill
                    status_cell.font = ok_font
                else:
                    status_cell.fill = warning_fill
                
                ws_card.cell(row=row, column=8, value=item['Fee_Applied']).border = border
                row += 1
            row += 1
        
        row += 1
        
        # ===== OUT SECTION =====
        ws_card.merge_cells(f'A{row}:H{row}')
        ws_card[f'A{row}'] = "📤 OUT TRANSACTIONS (Withdrawals/Cashouts)"
        ws_card[f'A{row}'].font = Font(bold=True, size=14, color="7030A0")
        ws_card[f'A{row}'].fill = gray_fill
        row += 2
        
        r_out = results['out']
        
        # OUT Summary
        ws_card[f'A{row}'] = "Card Journal:"
        ws_card[f'B{row}'] = r_out['summary']['card_count']
        ws_card[f'C{row}'] = "P2P Statement:"
        ws_card[f'D{row}'] = r_out['summary']['p2p_count']
        ws_card[f'E{row}'] = "Matched:"
        ws_card[f'F{row}'] = r_out['summary']['matched']
        ws_card[f'A{row}'].font = Font(bold=True)
        ws_card[f'C{row}'].font = Font(bold=True)
        ws_card[f'E{row}'].font = Font(bold=True)
        row += 2
        
        # Missing in P2P
        if r_out['missing_in_p2p']:
            ws_card[f'A{row}'] = "❌ MISSING IN P2P STATEMENT"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="C00000")
            row += 1
            
            headers = ['Account', 'Date', 'Time', 'Amount', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_red
                cell.border = border
            row += 1
            
            for item in r_out['missing_in_p2p']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['Time']).border = border
                cell_amt = ws_card.cell(row=row, column=4, value=item['Amount'])
                cell_amt.border = border
                cell_amt.number_format = '#,##0.00'
                ws_card.cell(row=row, column=5, value=item['Description']).border = border
                for c in range(1, 6):
                    ws_card.cell(row=row, column=c).fill = error_fill
                row += 1
            
            ws_card[f'A{row}'] = "Total Missing:"
            ws_card[f'A{row}'].font = Font(bold=True)
            ws_card.cell(row=row, column=4, value=r_out['missing_amount'])
            ws_card.cell(row=row, column=4).font = Font(bold=True)
            ws_card.cell(row=row, column=4).number_format = '#,##0.00'
            row += 2
        
        # Missing in Card
        if r_out['missing_in_card']:
            ws_card[f'A{row}'] = "❌ MISSING IN CARD JOURNAL"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="C00000")
            row += 1
            
            headers = ['Account', 'Date', 'Time', 'Amount', 'Type']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_red
                cell.border = border
            row += 1
            
            for item in r_out['missing_in_card']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['Time']).border = border
                cell_amt = ws_card.cell(row=row, column=4, value=item['Amount'])
                cell_amt.border = border
                cell_amt.number_format = '#,##0.00'
                ws_card.cell(row=row, column=5, value=item.get('Type', '-')).border = border
                for c in range(1, 6):
                    ws_card.cell(row=row, column=c).fill = error_fill
                row += 1
            row += 1
        
        # Matched OUT
        if r_out['matched_details']:
            ws_card[f'A{row}'] = "✓ MATCHED TRANSACTIONS"
            ws_card[f'A{row}'].font = Font(bold=True, size=11, color="006100")
            row += 1
            
            headers = ['Account', 'Date', 'P2P Amount', 'Adjusted', 'Card Amount', 'Diff', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws_card.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill_green
                cell.border = border
            row += 1
            
            for item in r_out['matched_details']:
                ws_card.cell(row=row, column=1, value=item['Auth']).border = border
                ws_card.cell(row=row, column=2, value=item['Date']).border = border
                ws_card.cell(row=row, column=3, value=item['P2P_Amount']).border = border
                ws_card.cell(row=row, column=4, value=item['Adjusted_Amount']).border = border
                ws_card.cell(row=row, column=5, value=item['Card_Amount']).border = border
                ws_card.cell(row=row, column=6, value=item['Difference']).border = border
                
                status_cell = ws_card.cell(row=row, column=7, value=item['Status'])
                status_cell.border = border
                if item['Status'] == 'OK':
                    status_cell.fill = ok_fill
                    status_cell.font = ok_font
                else:
                    status_cell.fill = warning_fill
                
                row += 1
        
        # Column widths for card sheet
        ws_card.column_dimensions['A'].width = 14
        ws_card.column_dimensions['B'].width = 12
        ws_card.column_dimensions['C'].width = 14
        ws_card.column_dimensions['D'].width = 12
        ws_card.column_dimensions['E'].width = 14
        ws_card.column_dimensions['F'].width = 10
        ws_card.column_dimensions['G'].width = 10
        ws_card.column_dimensions['H'].width = 35
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============== MAIN APP ==============
st.markdown('<div class="main-header">💳 Card Reconciliation Tool</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Compare Card Journal vs P2P Statement</div>', unsafe_allow_html=True)

st.info("""
**Rules:**
- 📥 **IN (Deposits):** Match by Auth code | 1% fee removed from P2P amounts > 40
- 📤 **OUT:** Match by Date + Amount + Account | No fee adjustment
- Use checkboxes to include/exclude: Withdrawal (Transfert vers), Mandat, Cashout
""")

# File upload section
st.markdown("---")
st.subheader("📁 Upload Files")

col1, col2 = st.columns(2)

with col1:
    st.markdown("**Card Journal**")
    st.caption("Files starting with card number (e.g., 178-D17-xxx.csv, 342-journal-xxx.csv)")
    card_files = st.file_uploader(
        "Upload Card Journal CSV files",
        type=['csv'],
        accept_multiple_files=True,
        key="card_journals"
    )

with col2:
    st.markdown("**P2P Statement**")
    st.caption("Files with statement-d17-XXX (e.g., statement-d17-178-xxx.csv)")
    p2p_files = st.file_uploader(
        "Upload P2P Statement CSV files",
        type=['csv'],
        accept_multiple_files=True,
        key="p2p_statements"
    )

# Date input
st.markdown("---")
col_date, col_options = st.columns([1, 2])

with col_date:
    from_date = st.date_input(
        "📅 Starting Date",
        value=datetime(2026, 1, 27),
        format="DD/MM/YYYY"
    )

with col_options:
    st.markdown("**OUT Transaction Options:**")
    col_opt1, col_opt2, col_opt3 = st.columns(3)
    with col_opt1:
        include_withdrawal = st.checkbox("Include Withdrawal", value=True)
    with col_opt2:
        include_mandat = st.checkbox("Include Mandat", value=True)
    with col_opt3:
        include_cashout = st.checkbox("Include Cashout", value=True)

# Process files
if card_files and p2p_files:
    # Organize files by card number
    card_journals = {}
    p2p_statements = {}
    
    for f in card_files:
        card_num, file_type = extract_card_number(f.name)
        if card_num and file_type == 'card_journal':
            card_journals[card_num] = f
    
    for f in p2p_files:
        card_num, file_type = extract_card_number(f.name)
        if card_num and file_type == 'p2p_statement':
            p2p_statements[card_num] = f
    
    # Show detected files
    st.markdown("---")
    st.subheader("📄 Detected Files")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Card Journals:**")
        for card, f in sorted(card_journals.items()):
            st.write(f"  Card {card}: {f.name}")
    
    with col2:
        st.markdown("**P2P Statements:**")
        for card, f in sorted(p2p_statements.items()):
            st.write(f"  Card {card}: {f.name}")
    
    # Find matched cards
    matched_cards = set(card_journals.keys()) & set(p2p_statements.keys())
    
    if matched_cards:
        st.success(f"✅ Matched cards: {', '.join(sorted(matched_cards))}")
        
        # Run reconciliation button
        if st.button("🚀 Run Reconciliation", type="primary", use_container_width=True):
            from_date_dt = datetime.combine(from_date, datetime.min.time())
            
            all_results = {}
            
            with st.spinner("Running reconciliation..."):
                for card_num in sorted(matched_cards):
                    try:
                        card_file = card_journals[card_num]
                        p2p_file = p2p_statements[card_num]
                        
                        card_file.seek(0)
                        p2p_file.seek(0)
                        
                        card_df = load_card_journal(card_file)
                        p2p_df = load_p2p_statement(p2p_file)
                        
                        # Get IN transactions
                        card_in = get_in_transactions_card(card_df, from_date_dt)
                        p2p_in = get_in_transactions_p2p(p2p_df, from_date_dt)
                        results_in = reconcile_transactions(card_in, p2p_in, 'IN')
                        
                        # Get OUT transactions
                        card_out = get_out_transactions_card(card_df, from_date_dt, include_withdrawal, include_mandat)
                        p2p_out = get_out_transactions_p2p(p2p_df, from_date_dt, include_withdrawal, include_cashout)
                        results_out = reconcile_transactions(card_out, p2p_out, 'OUT')
                        
                        all_results[card_num] = {
                            'in': results_in,
                            'out': results_out
                        }
                    except Exception as e:
                        st.error(f"Error processing card {card_num}: {e}")
            
            if all_results:
                # ==================== RECONCILIATION REPORT ====================
                st.markdown("---")
                
                # Header
                st.markdown("## 📊 RECONCILIATION REPORT")
                st.markdown(f"**Generated:** {datetime.now().strftime('%d/%m/%Y %H:%M')} | **Period From:** {from_date.strftime('%d/%m/%Y')}")
                
                # Options used
                options_text = []
                if include_withdrawal:
                    options_text.append("✓ Withdrawal")
                else:
                    options_text.append("✗ Withdrawal")
                if include_mandat:
                    options_text.append("✓ Mandat")
                else:
                    options_text.append("✗ Mandat")
                if include_cashout:
                    options_text.append("✓ Cashout")
                else:
                    options_text.append("✗ Cashout")
                st.caption(f"OUT Options: {' | '.join(options_text)}")
                
                # Calculate totals
                total_cards = len(all_results)
                total_in_matched = sum(r['in']['summary']['matched'] for r in all_results.values())
                total_in_missing_p2p = sum(r['in']['summary']['missing_in_p2p'] for r in all_results.values())
                total_in_missing_card = sum(r['in']['summary']['missing_in_card'] for r in all_results.values())
                total_in_missing_amount = sum(r['in']['missing_amount'] for r in all_results.values())
                
                total_out_matched = sum(r['out']['summary']['matched'] for r in all_results.values())
                total_out_missing_p2p = sum(r['out']['summary']['missing_in_p2p'] for r in all_results.values())
                total_out_missing_card = sum(r['out']['summary']['missing_in_card'] for r in all_results.values())
                total_out_missing_amount = sum(r['out']['missing_amount'] for r in all_results.values())
                
                total_missing = total_in_missing_amount + total_out_missing_amount
                
                # Overall Status
                st.markdown("### 🎯 Overall Status")
                if total_in_missing_p2p == 0 and total_in_missing_card == 0 and total_out_missing_p2p == 0 and total_out_missing_card == 0:
                    st.success("✅ **ALL TRANSACTIONS MATCHED** - No discrepancies found!")
                else:
                    st.error(f"⚠️ **DISCREPANCIES FOUND** - Total Missing Amount: **{total_missing:.2f}**")
                
                # Summary metrics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("📇 Cards Processed", total_cards)
                with col2:
                    delta_in = f"-{total_in_missing_p2p + total_in_missing_card}" if (total_in_missing_p2p + total_in_missing_card) > 0 else None
                    st.metric("📥 IN Matched", total_in_matched, delta=delta_in, delta_color="inverse")
                with col3:
                    delta_out = f"-{total_out_missing_p2p + total_out_missing_card}" if (total_out_missing_p2p + total_out_missing_card) > 0 else None
                    st.metric("📤 OUT Matched", total_out_matched, delta=delta_out, delta_color="inverse")
                with col4:
                    st.metric("💰 Missing Amount", f"{total_missing:.2f}")
                
                st.markdown("---")
                
                # ==================== IN TRANSACTIONS SUMMARY ====================
                st.markdown("### 📥 IN Transactions (Deposits)")
                in_data = []
                for card_num, results in sorted(all_results.items()):
                    r = results['in']
                    summary = r['summary']
                    status = "✅" if summary['missing_in_p2p'] == 0 and summary['missing_in_card'] == 0 else "❌"
                    in_data.append({
                        'Status': status,
                        'Card': card_num,
                        'Card Journal': summary['card_count'],
                        'P2P Statement': summary['p2p_count'],
                        'Matched': summary['matched'],
                        'Missing (P2P)': summary['missing_in_p2p'],
                        'Missing (Card)': summary['missing_in_card'],
                        'Missing Amount': f"{r['missing_amount']:.2f}" if r['missing_amount'] > 0 else "-"
                    })
                
                df_in = pd.DataFrame(in_data)
                st.dataframe(df_in, use_container_width=True, hide_index=True)
                
                if total_in_missing_amount > 0:
                    st.error(f"⚠️ Total IN Missing: **{total_in_missing_amount:.2f}**")
                else:
                    st.success("✅ All IN transactions matched!")
                
                # ==================== OUT TRANSACTIONS SUMMARY ====================
                st.markdown("### 📤 OUT Transactions (Withdrawals)")
                out_data = []
                for card_num, results in sorted(all_results.items()):
                    r = results['out']
                    summary = r['summary']
                    status = "✅" if summary['missing_in_p2p'] == 0 and summary['missing_in_card'] == 0 else "❌"
                    out_data.append({
                        'Status': status,
                        'Card': card_num,
                        'Card Journal': summary['card_count'],
                        'P2P Statement': summary['p2p_count'],
                        'Matched': summary['matched'],
                        'Missing (P2P)': summary['missing_in_p2p'],
                        'Missing (Card)': summary['missing_in_card'],
                        'Missing Amount': f"{r['missing_amount']:.2f}" if r['missing_amount'] > 0 else "-"
                    })
                
                df_out = pd.DataFrame(out_data)
                st.dataframe(df_out, use_container_width=True, hide_index=True)
                
                if total_out_missing_amount > 0:
                    st.error(f"⚠️ Total OUT Missing: **{total_out_missing_amount:.2f}**")
                else:
                    st.success("✅ All OUT transactions matched!")
                
                # ==================== MISSING TRANSACTIONS DETAIL ====================
                # Collect all missing transactions
                all_missing_in_p2p = []
                all_missing_in_card = []
                all_missing_out_p2p = []
                all_missing_out_card = []
                
                for card_num, results in sorted(all_results.items()):
                    for item in results['in']['missing_in_p2p']:
                        item_copy = item.copy()
                        item_copy['Card'] = card_num
                        all_missing_in_p2p.append(item_copy)
                    for item in results['in']['missing_in_card']:
                        item_copy = item.copy()
                        item_copy['Card'] = card_num
                        all_missing_in_card.append(item_copy)
                    for item in results['out']['missing_in_p2p']:
                        item_copy = item.copy()
                        item_copy['Card'] = card_num
                        all_missing_out_p2p.append(item_copy)
                    for item in results['out']['missing_in_card']:
                        item_copy = item.copy()
                        item_copy['Card'] = card_num
                        all_missing_out_card.append(item_copy)
                
                if all_missing_in_p2p or all_missing_in_card or all_missing_out_p2p or all_missing_out_card:
                    st.markdown("---")
                    st.markdown("### ❌ MISSING TRANSACTIONS DETAIL")
                    
                    if all_missing_in_p2p:
                        st.markdown("#### 📥 IN - Missing in P2P Statement")
                        df = pd.DataFrame(all_missing_in_p2p)
                        cols = ['Card'] + [c for c in df.columns if c != 'Card']
                        st.dataframe(df[cols], use_container_width=True, hide_index=True)
                    
                    if all_missing_in_card:
                        st.markdown("#### 📥 IN - Missing in Card Journal")
                        df = pd.DataFrame(all_missing_in_card)
                        cols = ['Card'] + [c for c in df.columns if c != 'Card']
                        st.dataframe(df[cols], use_container_width=True, hide_index=True)
                    
                    if all_missing_out_p2p:
                        st.markdown("#### 📤 OUT - Missing in P2P Statement")
                        df = pd.DataFrame(all_missing_out_p2p)
                        cols = ['Card'] + [c for c in df.columns if c != 'Card']
                        st.dataframe(df[cols], use_container_width=True, hide_index=True)
                    
                    if all_missing_out_card:
                        st.markdown("#### 📤 OUT - Missing in Card Journal")
                        df = pd.DataFrame(all_missing_out_card)
                        cols = ['Card'] + [c for c in df.columns if c != 'Card']
                        st.dataframe(df[cols], use_container_width=True, hide_index=True)
                
                # ==================== CARD DETAILS (Expandable) ====================
                st.markdown("---")
                st.markdown("### 📋 Detailed View by Card")
                
                for card_num, results in sorted(all_results.items()):
                    in_status = "✅" if results['in']['summary']['missing_in_p2p'] == 0 and results['in']['summary']['missing_in_card'] == 0 else "❌"
                    out_status = "✅" if results['out']['summary']['missing_in_p2p'] == 0 and results['out']['summary']['missing_in_card'] == 0 else "❌"
                    
                    with st.expander(f"Card {card_num} | IN: {in_status} ({results['in']['summary']['matched']} matched) | OUT: {out_status} ({results['out']['summary']['matched']} matched)"):
                        tab_in, tab_out = st.tabs(["📥 IN", "📤 OUT"])
                        
                        with tab_in:
                            r = results['in']
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Matched", r['summary']['matched'])
                            with col2:
                                st.metric("Missing P2P", r['summary']['missing_in_p2p'])
                            with col3:
                                st.metric("Missing Card", r['summary']['missing_in_card'])
                            
                            if r['missing_in_p2p']:
                                st.markdown("**❌ Missing in P2P Statement:**")
                                st.dataframe(pd.DataFrame(r['missing_in_p2p']), use_container_width=True, hide_index=True)
                            
                            if r['missing_in_card']:
                                st.markdown("**❌ Missing in Card Journal:**")
                                st.dataframe(pd.DataFrame(r['missing_in_card']), use_container_width=True, hide_index=True)
                            
                            if r['matched_details']:
                                st.markdown("**✅ Matched Transactions:**")
                                st.dataframe(pd.DataFrame(r['matched_details']), use_container_width=True, hide_index=True)
                        
                        with tab_out:
                            r = results['out']
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Matched", r['summary']['matched'])
                            with col2:
                                st.metric("Missing P2P", r['summary']['missing_in_p2p'])
                            with col3:
                                st.metric("Missing Card", r['summary']['missing_in_card'])
                            
                            if r['missing_in_p2p']:
                                st.markdown("**❌ Missing in P2P Statement:**")
                                st.dataframe(pd.DataFrame(r['missing_in_p2p']), use_container_width=True, hide_index=True)
                            
                            if r['missing_in_card']:
                                st.markdown("**❌ Missing in Card Journal:**")
                                st.dataframe(pd.DataFrame(r['missing_in_card']), use_container_width=True, hide_index=True)
                            
                            if r['matched_details']:
                                st.markdown("**✅ Matched Transactions:**")
                                st.dataframe(pd.DataFrame(r['matched_details']), use_container_width=True, hide_index=True)
                
                # ==================== DOWNLOAD REPORT ====================
                st.markdown("---")
                excel_file = create_excel_report(all_results, from_date_dt, include_withdrawal, include_mandat, include_cashout)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_file,
                    file_name=f"recon_report_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
    else:
        st.warning("⚠️ No matching card pairs found. Make sure you have both Card Journal and P2P Statement files for at least one card.")

else:
    st.markdown("---")
    st.markdown("### 👆 Upload your files to get started")
    st.markdown("""
    **File naming (flexible):**
    - Card Journal: Any file starting with card number (e.g., `178-D17-xxx.csv`, `342-journal-xxx.csv`)
    - P2P Statement: Files containing `statement-d17-XXX` (e.g., `statement-d17-178-xxx.csv`)
    """)

# Footer
st.markdown("---")
st.caption("Card Reconciliation Tool v2.8 | Withdrawal/Mandat/Cashout Options")
